from tkinter import *
from PIL import ImageTk, Image 
from tkinter import messagebox, ttk
from tkcalendar import Calendar, DateEntry
from tkinter.filedialog import askopenfilename
import openpyxl ,xlrd
from openpyxl import Workbook
import pathlib
import re
import random
import smtplib
from email.message import EmailMessage

rip =Tk()
rip.geometry('1500x1500')
rip.title("School Registration Form")
rip.state('zoomed')

nw_color='Green'

#images/////
login_stud_icn=PhotoImage(file=r'login_student_img.png')
admi_icn=PhotoImage(file=r'admin_img.png')
add_icn=PhotoImage(file=r'add_student_img.png')
add_image=PhotoImage(file=r'add_image.png')
bg_image=PhotoImage(file=r'Screenshot (29).png')
# bg_image1=PhotoImage(file=r"Screenshot (17).png")
bg_image2=PhotoImage(file=r"Screenshot (24).png")
bg_image4=PhotoImage(file=r"Screenshot (27).png")
bg_image5=PhotoImage(file=r"Screenshot (28).png")
bg_image6=PhotoImage(file=r"Screenshot (30).png")

file=pathlib.Path('Back.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Student Name"
    sheet['B1']="DOB"
    sheet['C1']="Adress"
    sheet['D1']="Father Name"
    sheet['E1']="Mother Name"
    sheet['F1']="Email"
    sheet['G1']="Student ID"


    file.save('Back.xlsx')

li=[]

def welcome_page():
    # Image add
    image5 = Label(image=bg_image5)
    image5.place(x=0, y=0, anchor="nw")


    def destory():
        Welcome.destroy()
        image5.destroy()
        rip.update()
        student_login()

    def admin():
        Welcome.destroy()
        image5.destroy()        
        rip.update()
        student_info() 

    def student():
        Welcome.destroy()
        image5.destroy()        
        rip.update()
        admin_login()


    
    Welcome=Frame(rip,highlightbackground=nw_color,highlightthickness=3)

    border=Label(Welcome,text="Welcome",bg=nw_color,fg="white",font=('bold',18))
    border.place(x=0,y=0,width=400)

    #...student button
    stud=Button(Welcome,text="Student Login",bg=nw_color,fg="white",font=('bold',15),bd=0,command=destory)
    stud.place(x=120,y=125,width=200)
    #student icon
    stud_icon=Button(Welcome,image=login_stud_icn,bd=0,command=destory)
    stud_icon.place(x=60,y=100)

    #...admin button
    stud=Button(Welcome,text="Admin Login",bg=nw_color,fg="white",font=('bold',15),bd=0,command=student)
    stud.place(x=120,y=225,width=200)
    #admin icon
    admi_icon=Button(Welcome,image=admi_icn,bd=0,command=student)
    admi_icon.place(x=60,y=200)

    #...create account button
    stud=Button(Welcome,text="Create Account",bg=nw_color,fg="white",font=('bold',15),bd=0,command=admin)
    stud.place(x=120,y=325,width=200)
    #student icon
    stud_icon=Button(Welcome,image=login_stud_icn,bd=0,command=admin)
    stud_icon.place(x=60,y=300)

    Welcome.pack(pady=30)
    Welcome.pack_propagate(False)
    Welcome.configure(width=400,height=420)

def student_login():
    def framedes():
        frame.destroy()
        image.destroy()
        rip.update()
        welcome_page()

    #submitfunction
    b1=StringVar()
    b2=StringVar()

    def submit():
        print("Submit button pressed")
        a = b1.get()
        b = b2.get()

        print("Values of a and b:", a, b)

        if a.strip() == "" or b.strip() == "":
            messagebox.showerror("Error", "Please fill all fields")
        elif a == li[13] and b == li[0]:
            messagebox.showinfo("Login successful")
            student_ID_card()
        else:
            messagebox.showerror("Error", "Incorrect username or password")



    def password_visibility():
      
        if show_password_var.get():
            label_PasswordE.config(show="")
        else:
            label_PasswordE.config(show="*")

    # Image add
    image = Label(image=bg_image)
    image.place(x=0, y=0, anchor="nw")

    # Frame
    frame = Frame(rip, width=350, height=450, bg='white')
    frame.place(relx=0.4, rely=0.55, anchor="center")

    # transparent white box 
    # frame_bg = Label(frame, borderwidth=0)
    # frame_bg.place(relx=0.5, rely=0.5, anchor="center")

    #loginLogo
    login=Label(frame,text="Student Login",font=("Elephant",19,"italic","underline"),bg="white",fg="firebrick1")
    login.place(relx=0.5, rely=0.15, anchor="center",)

    # Username Field
    label_username = Label(frame, text="Student ID", bg="white", fg="firebrick1", font=("Arial", 14, "italic","bold",))
    label_usernameE = Entry(frame, font=("Arial", 12, "italic"),textvariable=b1)
    label_username.place(relx=0.18, rely=0.4, anchor="center")
    label_usernameE.place(relx=0.6, rely=0.4, anchor="center", width=140)

    # Password Field
    label_Password = Label(frame, text="Stu. Name", bg="White", fg="firebrick1", font=("Arial", 14, "italic","bold"))
    label_PasswordE = Entry(frame,show="****",font=("Arial", 12, "italic"),textvariable=b2)
    label_Password.place(relx=0.18, rely=0.6, anchor="center")
    label_PasswordE.place(relx=0.6, rely=0.6, anchor="center", width=140)

    # Show Password Checkbox
    show_password_var = IntVar()
    show_password_checkbox = Checkbutton(frame, variable=show_password_var, command=password_visibility,bg="white",fg="black")
    show_password_checkbox.place(relx=0.9, rely=0.6, anchor="center")


    # Button
    login_button = Button(frame, text="Login", font=("Arial", 14,"bold"), bg="#4CAF50", fg="white",command=submit)
    login_button.place(relx=0.55, rely=0.78, anchor="center", relwidth=0.4)

    # # Signup
    # signup_label = Label(frame, text="Don't have an account?", font=("Arial", 12,"italic"),bg="white")
    # signup_label.place(relx=0.34, rely=0.9, anchor="center")

    # # Signup Button
    # signup_button = Button(frame, text="Sign Up", font=("Arial", 11, "italic"), fg="blue",bg="white",bd=0)
    # signup_button.place(relx=0.75, rely=0.9, anchor='center')

    #next button
    next_button3=Button(frame,text="Back",bg="firebrick1",fg="white",bd=0,font=("Arial",16,"bold"),padx=4,command=framedes)
    next_button3.place(relx=0.01,rely=0.01)

def admin_login():
    def frame1des():
        frame1.destroy()
        image.destroy()
        rip.update()
        welcome_page()

    #submitfunction
    b1=StringVar()
    b2=StringVar()

    def submit():
        a=b1.get()
        b=b2.get()
        b1.set("")
        b2.set("")


        if a=="" or b=="":
            messagebox.showerror("Error","Please fill all Fields")
        else:
            messagebox.showinfo("Done","Data is Submitted Successfully")
            print("Username :",a)
            print("Password :",b)



    def password_visibility():
      
        if show_password_var.get():
            label_PasswordE.config(show="")
        else:
            label_PasswordE.config(show="*")


    # Image add
    image = Label(image=bg_image)
    image.place(x=0, y=0, anchor="nw")

    # Frame
    frame1 = Frame(rip, width=350, height=450, bg='white')
    frame1.place(relx=0.4, rely=0.55, anchor="center")

    # transparent white box 
    # frame_bg = Label(frame, borderwidth=0)
    # frame_bg.place(relx=0.5, rely=0.5, anchor="center")

    #loginLogo
    login=Label(frame1,text="Admin Login",font=("Elephant",19,"italic","underline"),bg="white",fg="firebrick1")
    login.place(relx=0.5, rely=0.15, anchor="center",)

    # Username Field
    label_username = Label(frame1, text="Admin Name", bg="white", fg="firebrick1", font=("Arial", 14, "italic","bold"))
    label_usernameE = Entry(frame1, font=("Arial", 12, "italic"),textvariable=b1)
    label_username.place(relx=0.2, rely=0.4, anchor="center")
    label_usernameE.place(relx=0.65, rely=0.4, anchor="center", width=140)

    # Password Field
    label_Password = Label(frame1, text="Password", bg="White", fg="firebrick1", font=("Arial", 14, "italic","bold"))
    label_PasswordE = Entry(frame1,show="",font=("Arial", 12, "italic"),textvariable=b2)
    label_Password.place(relx=0.18, rely=0.6, anchor="center",relwidth=0.3)
    label_PasswordE.place(relx=0.65, rely=0.6, anchor="center", width=140)

    # Show Password Checkbox
    show_password_var = IntVar()
    show_password_checkbox = Checkbutton(frame1, variable=show_password_var, command=password_visibility,bg="white",fg="black")
    show_password_checkbox.place(relx=0.95, rely=0.6, anchor="center")


    # Button
    login_button = Button(frame1, text="Login", font=("Arial", 14,"bold"), bg="#4CAF50", fg="white",command=submit)
    login_button.place(relx=0.55, rely=0.78, anchor="center", relwidth=0.4)

    #next button
    next_button4=Button(frame1,text="Back",bg="firebrick1",fg="white",bd=0,font=("Arial",16,"bold"),padx=10,command=frame1des)
    next_button4.place(relx=0.01,rely=0.01)


def student_info():

    def validate_student_info():
        
           # Validation for each field
        student_name = student_name_entry.get()
        dob = cal.get()
        gender = gender_var.get()
        address = address_entry.get()
        zip_code = zip_entry.get()
        state = state_entry.get()


        if not student_name:
            messagebox.showerror("Error", "Please enter Student Name.")
            return False
        if not dob:
            messagebox.showerror("Error", "Please select Date of Birth.")
            return False
        if not gender:
            messagebox.showerror("Error", "Please select Gender.")
            return False
        if not address:
            messagebox.showerror("Error", "Please enter Address.")
            return False
        if not zip_code.isdigit() or zip_code == "":
            messagebox.showerror("Error", "Please enter Numeric Character\nin Zip Code")
            return False
        if not zip_code:
            messagebox.showerror("Error", "Please enter zip_code")
            return False        
        if not state:
            messagebox.showerror("Error", "Please enter State.")
            return False

        li.append(student_name)
        li.append(dob)
        li.append(gender)
        li.append(address)
        li.append(zip_code)
        li.append(state)

        # Print entered data
        print("Student Name:", student_name)
        print("Date of Birth:", dob)
        print("Gender:", gender)
        print("Address:", address)
        print("Zip Code:", zip_code)
        print("State:", state)
        file=openpyxl.load_workbook('Back.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=student_name)
        sheet.cell(column=2,row=sheet.max_row,value=dob)
        sheet.cell(column=3,row=sheet.max_row,value=address)
        file.save('Back.xlsx')
        
        return True

    def frame2dest():
        if validate_student_info():
            frame2.destroy()
            image1.destroy()
            rip.update()
            parents_info()

    def forwelcome():
        frame2.destroy()
        image1.destroy()
        rip.update()
        welcome_page()
    # Image add
    image1 = Label(image=bg_image2)
    image1.place(x=0, y=0, anchor="nw")


    frame2 = Frame(rip, width=300, height=434, bg='white')
    frame2.pack(padx=20, pady=20)

    student_label=Label(frame2,text="Student Information",bg="white",fg="firebrick1",font=("Arial",15,"bold","underline"))
    student_label.place(relx=0.18,rely=0.02)

    #student name
    student_name=Label(frame2,text="Student Name :",bg="white",fg="black",font=("Arial",12,"bold"))
    student_name.place(relx=0.02,rely=0.15)
    student_name_entry=Entry(frame2,font=("Arial",12,"normal"),bd=0,width=12)
    student_name_entry.place(relx=0.45,rely=0.15)
    Frame(frame2,width=103,height=2,bg='firebrick1').place(relx=0.45,rely=0.2)

    #DOB
    DOB_label=Label(frame2,text="DOB",bg="white",fg="black",font=("Arial",12,"bold"))
    DOB_label.place(relx=0.02,rely=0.25)
    Frame(frame2,width=103,height=2,bg='firebrick1').place(relx=0.45,rely=0.3)
    cal = DateEntry(selectmode='day',bd=0,bg="white",width=13)
    cal.place(relx=0.49,rely=0.175)

    #Gender
    gender = Label(frame2,text="Gender :", bg="white", font=("Arial", 12, "bold"))
    gender.place(relx=0.129, rely=0.37, anchor="center")

    gender_var = StringVar(frame2, "Male")

    male_radio = Radiobutton(frame2, text="Male", variable=gender_var, value="Male", bg="white",fg="firebrick1",font=("Arial", 12, "bold"))
    male_radio.place(relx=0.5, rely=0.37, anchor="center")

    female_radio = Radiobutton(frame2, text="Female", variable=gender_var, value="Female", bg="white",fg="firebrick1",font=("Arial", 12, "bold"))
    female_radio.place(relx=0.787, rely=0.37, anchor="center")

    #adress
    address_label=Label(frame2,text="Address :",bg="white",fg="black",font=("Arial",12,"bold"))
    address_label.place(relx=0.014,rely=0.44)
    address_entry=Entry(frame2,font=("Arial",12,"normal"),bd=0,width=18)
    address_entry.place(relx=0.36,rely=0.428)
    Frame(frame2,width=170,height=2,bg='firebrick1').place(relx=0.36,rely=0.48)

    #zipcode
    zip_label=Label(frame2,text="Zip Code :",bg="white",fg="black",font=("Arial",12,"bold"))
    zip_label.place(relx=0.014,rely=0.53)
    zip_entry=Entry(frame2,font=("Arial",12,"normal"),bd=0,width=18)
    zip_entry.place(relx=0.4,rely=0.53)
    Frame(frame2,width=103,height=2,bg='firebrick1').place(relx=0.4,rely=0.58)

    #state
    state_label=Label(frame2,text="State :",bg="white",fg="black",font=("Arial",13,"bold"))
    state_label.place(relx=0.014,rely=0.64)
    state_entry=Entry(frame2,font=("Arial",12,"normal"),bd=0,width=18)
    state_entry.place(relx=0.4,rely=0.64)
    Frame(frame2,width=103,height=2,bg='firebrick1').place(relx=0.4,rely=0.68)

    next_button1=Button(frame2,text="Next",bg="blue",fg="white",font=("Arial",16,"bold"),padx=20,command=frame2dest)
    next_button1.place(relx=0.55,rely=0.8)

    exit_button1=Button(frame2,text="Back",bg="blue",fg="white",font=("Arial",16,"bold"),padx=20,command=forwelcome)
    exit_button1.place(relx=0.1,rely=0.8)

def parents_info():
    # Image add
    image2 = Label(image=bg_image2)
    image2.place(x=0, y=0, anchor="nw")

    def validate_parents_info():
        # Validation for each field
        father_name = father_name_entry.get()
        father_occupation = father_occupation_entry.get()
        mother_name = mother_name_entry.get()
        mother_occupation = mother_occupation_entry.get()
        father_phone = father_phone_entry.get()
        mother_phone = mother_phone_entry.get()

        if not father_name:
            messagebox.showerror("Error", "Please enter Father/Guardian's Name.")
            return False
        if not father_occupation:
            messagebox.showerror("Error", "Please enter Father's Occupation.")
            return False
        if not mother_name:
            messagebox.showerror("Error", "Please enter Mother's Name.")
            return False
        if not mother_occupation:
            messagebox.showerror("Error", "Please enter Mother's Occupation.")
            return False
        if not father_phone.isdigit() or father_phone == "" or len(father_phone) != 10:
            messagebox.showerror("Error", "Please enter Numeric Character only\nshould be 10 Digit in contact fi")
            return False
        if not mother_phone.isdigit() or mother_phone == ""  or len(mother_phone) != 10:
            messagebox.showerror("Error", "Please enter Numeric Character only\nshould be 10 Digitin contact field")
            return False

        li.append(father_name)
        li.append(father_occupation)
        li.append(mother_name)
        li.append(mother_occupation)
        li.append(father_phone)
        li.append(mother_phone)
    
        # Print entered data
        print("Father's Name:", father_name)
        print("Father's Occupation:", father_occupation)
        print("Mother's Name:", mother_name)
        print("Mother's Occupation:", mother_occupation)
        print("Father's Phone:", father_phone)
        print("Mother's Phone:", mother_phone)

        file=openpyxl.load_workbook('Back.xlsx')
        sheet=file.active
        sheet.cell(column=4,row=sheet.max_row,value=father_name)
        sheet.cell(column=5,row=sheet.max_row,value=mother_name)
        file.save('Back.xlsx')

        
        return True


    def back1():
        frame3.destroy()
        image2.destroy()
        rip.update()
        student_info()

    def next1():
        if validate_parents_info():
            frame3.destroy()
            image2.destroy()
            rip.update()
            verifi_cation()

    #Adding BGImage
    frame3 = Frame(rip, width=400, height=434, bg='white')
    frame3.pack(padx=20, pady=20)

    parents_label=Label(frame3,text="Parents/Guardians Information",bg="white",fg="firebrick1",font=("Arial",15,"bold","underline"))
    parents_label.place(relx=0.1,rely=0.02)

    #father name
    Father_name=Label(frame3,text="Father/Guardians Name :",bg="white",fg="black",font=("Arial",12,"bold"))
    Father_name.place(relx=0.02,rely=0.15)
    father_name_entry=Entry(frame3,font=("Arial",12,"normal"),bd=0,width=15)
    father_name_entry.place(relx=0.53,rely=0.15)
    Frame(frame3,width=140,height=2,bg='firebrick1').place(relx=0.53,rely=0.2)

    #FOcupation
    Foccupation_label=Label(frame3,text="Father Occupation :",bg="white",fg="black",font=("Arial",12,"bold"))
    Foccupation_label.place(relx=0.02,rely=0.25)
    father_occupation_entry=Entry(frame3,font=("Arial",12,"normal"),bd=0,width=15)
    father_occupation_entry.place(relx=0.53,rely=0.25)
    Frame(frame3,width=140,height=2,bg='firebrick1').place(relx=0.53,rely=0.3)

    #Mother name
    Mother_name=Label(frame3,text="Mother Name :",bg="white",fg="black",font=("Arial",12,"bold"))
    Mother_name.place(relx=0.02,rely=0.35)
    mother_name_entry=Entry(frame3,font=("Arial",12,"normal"),bd=0,width=15)
    mother_name_entry.place(relx=0.53,rely=0.35)
    Frame(frame3,width=140,height=2,bg='firebrick1').place(relx=0.53,rely=0.4)

    #mOcupation
    Moccupation_label=Label(frame3,text="Mother Occupation :",bg="white",fg="black",font=("Arial",12,"bold"))
    Moccupation_label.place(relx=0.02,rely=0.45)
    mother_occupation_entry=Entry(frame3,font=("Arial",12,"normal"),bd=0,width=15)
    mother_occupation_entry.place(relx=0.53,rely=0.45)
    Frame(frame3,width=140,height=2,bg='firebrick1').place(relx=0.53,rely=0.5)

    #father phoneno.
    Father_Phone=Label(frame3,text="Father/Guard contact no.:",bg="white",fg="black",font=("Arial",12,"bold"))
    Father_Phone.place(relx=0.02,rely=0.55)
    father_phone_entry=Entry(frame3,font=("Arial",12,"normal"),bd=0,width=15)
    father_phone_entry.place(relx=0.53,rely=0.55)
    Frame(frame3,width=140,height=2,bg='firebrick1').place(relx=0.53,rely=0.6)

    #mother phoneno.
    Mother_Phone=Label(frame3,text="Mother Contact no. :",bg="white",fg="black",font=("Arial",12,"bold"))
    Mother_Phone.place(relx=0.02,rely=0.65)
    mother_phone_entry=Entry(frame3,font=("Arial",12,"normal"),bd=0,width=15)
    mother_phone_entry.place(relx=0.53,rely=0.65)
    Frame(frame3,width=140,height=2,bg='firebrick1').place(relx=0.53,rely=0.7)

    #next button
    next_button=Button(frame3,text="Next",bg="blue",fg="white",font=("Arial",16,"bold"),padx=20,command=next1)
    next_button.place(relx=0.55,rely=0.8)

    exit_button=Button(frame3,text="Back",bg="blue",fg="white",font=("Arial",16,"bold"),padx=20,command=back1)
    exit_button.place(relx=0.1,rely=0.8)


def verifi_cation():
        # Image add
    image3 = Label(image=bg_image4)
    image3.place(x=0, y=0, anchor="nw")

    current_captcha = ""

    def generate_captcha():
        # Generate a random 4-character string for the CAPTCHA
        captcha_text = ''.join(random.choices('abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789', k=4))
        return captcha_text

    def draw_captcha(canvas, captcha_text):
        canvas.delete("captcha_text")  # Clear previous captcha text
        canvas.create_text(40, 18, text=captcha_text, font=('Arial', 13), tag="captcha_text")

    def validate_captcha():
        # Check if the entered captcha matches the generated one
        entered_captcha = captcha_entry.get()
        if entered_captcha == current_captcha:
            messagebox.showinfo("Success", "CAPTCHA matched!")
        else:
            messagebox.showerror("Error", "CAPTCHA didn't match. Please try again.")
            generate_and_draw_captcha()

    def generate_and_draw_captcha():
        nonlocal current_captcha
        current_captcha = generate_captcha()
        draw_captcha(canvas, current_captcha)

    pic_path = StringVar()
    pic_path.set('')

    def open_pic():
        path = askopenfilename()

        if path:
            img = Image.open(path).resize((100, 100))
            img3 = ImageTk.PhotoImage(img)
            pic_path.set(path)

            imgbutton.config(image=img3)
            imgbutton.image = img3

    def check_invalid_email(email):
        pattern = r"^[_a-z0-9-]+(\.[_a-z0-9-]+)*@[a-z0-9-]+(\.[a-z0-9-]+)*(\.[a-z]{2,4})$"
        match = re.match(pattern=pattern, string=email)
        return match

    def generate_id_number():
        genrated_id =''

        for r in range(6):

            genrated_id += str(random.randint(0, 9))

        print('id number: ', genrated_id)

        studentid_entry.config(state='normal')
        studentid_entry.delete(0,END)
        studentid_entry.insert(END,genrated_id)
        studentid_entry.config(state='readonly')


    def validate_and_store_data():
    # Validation for each field
        if pic_path.get() == '':
            messagebox.showerror("Error", "Please select an image.")
            return
        if not class_var.get():
            messagebox.showerror("Error", "Please select a class.")
            return
        if not medium_var.get():
            messagebox.showerror("Error", "Please select a medium.")
            return
        email = email_entry.get()
        if not email:
            messagebox.showerror("Error", "Please enter an email.")
            return
        if not check_invalid_email(email=email_entry.get().lower()):
            messagebox.showerror("Error", "Please enter a valid email address.")
            return

        if otp != verify_otp_entry.get():
            messagebox.showerror("Error", "Please verify Gmail with OTP.")
            return 

        # Check if captcha is verified
        if current_captcha != captcha_entry.get():
            messagebox.showerror("Error", "Please verify the captcha.")
            return

         
        # Store the data (you can modify this part according to your requirements)
        selected_image_path = pic_path.get()
        selected_class = class_var.get()
        selected_medium = medium_var.get()
        entered_email = email
        st_id=studentid_entry.get()

        # Here you can store the data as required, such as in a database or file.
        # For demonstration purposes, I'm just printing the data.
        print("Selected Image Path:", selected_image_path)
        print("Selected Class:", selected_class)
        print("Selected Medium:", selected_medium)
        print("Entered Email:", entered_email)
        print("Student Id:", st_id)

        li.append(selected_image_path)
        li.append(st_id)
        li.append(selected_class)
        li.append(selected_medium)
        li.append(entered_email)
        print(li)

        file=openpyxl.load_workbook('Back.xlsx')
        sheet=file.active
        sheet.cell(column=6,row=sheet.max_row,value=entered_email)
        sheet.cell(column=7,row=sheet.max_row,value=st_id)
        file.save('Back.xlsx')


        # Display a message to the user
        messagebox.showinfo("Success", "Form submitted successfully.")

        # Clear all form fields
        pic_path.set('')  # Clear the image path variable
        class_var.set('')  # Clear the class variable
        medium_var.set('')  # Clear the medium variable
        email_entry.delete(0, 'end')  # Clear the email entry field
        frame4.destroy()
        image3.destroy()
        rip.update()
        student_login()#>>>>>>>>>>>>>>>>>>>>>>>>>>>CHANGE IT


    def back2(): 
        frame4.destroy()
        image3.destroy()
        rip.update()
        parents_info()

    #Adding BGImage
    frame4 = Frame(rip, width=350, height=500, bg='white')
    frame4.pack(padx=20, pady=20)

    #adding frame for image
    imgf = Frame(frame4, width=105, height=105, highlightbackground="red", highlightthickness=2)
    imgbutton = Button(frame4, image=add_image, command=open_pic)
    imgbutton.place(relx=0.359, rely=0.025)
    imgf.place(relx=0.356, rely=0.02)

    # Create a canvas to draw the CAPTCHA
    canvas = Canvas(frame4, width=80, height=28, bg='white')
    canvas.place(relx=0.05, rely=0.72)

    # Generate and draw the initial CAPTCHA
    generate_and_draw_captcha()

    # Captcha Entry
    captcha_entry = Entry(frame4, font=("Arial", 12, "normal"), width=10)
    captcha_entry.place(relx=0.05, rely=0.8)

    # Student Id
    studentid_label = Label(frame4, text="Student ID:", bg="white", fg="black", font=("Arial", 11, "bold"))
    idremember = Label(frame4, text="*Remember Your Id", bg="white", fg="red", font=("Arial", 8, "bold")).place(relx=0.35,rely=0.31)
    studentid_label.place(relx=0.05, rely=0.27)

    # Student ID entry
    studentid_entry = Entry(frame4, bg="white", fg="black", font=("Arial", 13, "bold"), width=15, bd=0)
    studentid_entry.place(relx=0.35, rely=0.27)
    generate_id_number()
    Frame(frame4, width=150, height=2, bg='firebrick1').place(relx=0.35, rely=0.31)

    # Set entry to readonly
    studentid_entry.config(state='readonly')


    #class_label
    class_label = Label(frame4, text="Class:", bg="white", fg="black", font=("Arial", 11, "bold"))
    class_label.place(relx=0.05, rely=0.37)
    classes = ["Class 1", "Class 2", "Class 3", "Class 4", "Class 5", "Class 6", "class 7", "Class 8", "Class 9", "Class 10"]
    class_var = StringVar(frame4)
    class_var.set(classes[0])
    OptionMenu(frame4, class_var, *classes).place(relx=0.53, rely=0.37)

    #medium 
    medium = Label(frame4, text="Medium:", bg="white", fg="black", font=("Arial", 11, "bold"))
    medium.place(relx=0.05, rely=0.46)
    mediums = ["English", "Hindi"]
    medium_var = StringVar(frame4)
    medium_var.set(mediums[0])
    OptionMenu(frame4, medium_var, *mediums).place(relx=0.53, rely=0.46)

    #Email
    Label(frame4, text="Email:", bg="white", fg="black", font=("Arial", 11, "bold")).place(relx=0.05, rely=0.56)
    email_entry = Entry(frame4, font=("Arial", 12, "normal"), width=22, bd=0)
    email_entry.place(relx=0.245, rely=0.56)
    Frame(frame4, width=195, height=2, bg='firebrick1').place(relx=0.245, rely=0.6)
    #otp>>>>>>>>>>>>>.............................................................................................
    #sentotp
    otp = ""
    for i in range(6):
        otp += str(random.randint(0,9))

    print(otp)

    def sendmail():
        # Assuming check_invalid_email function is defined elsewhere in your code
        if not check_invalid_email(email=email_entry.get().lower()):
            messagebox.showerror("Error", "Please enter a valid email address.")
        else:
            try:
                # Mail setup
                server = smtplib.SMTP("smtp.gmail.com", 587)
                server.starttls()
                from_mail = "rudragill9871@gmail.com"
                app_password = "wbcx nmnj cygm lrza"
                server.login(from_mail, app_password)

                user_mail = email_entry.get()
                Emsg = EmailMessage()
                Emsg["Subject"] = "OTP VERIFICATION"
                Emsg["From"] = from_mail
                Emsg["To"] = user_mail
                Emsg.set_content("Your OTP For Verification is " +otp+ ".  Don't Share With Any One")

                server.send_message(Emsg)
                server.quit()

                messagebox.showinfo("Information", "Email is sent to your mail")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to send email: {str(e)}")

    def verify_otp():
        # Get the OTP entered by the user
        entered_otp = verify_otp_entry.get()  # Access the value from the entry widget
        
        # Compare with the OTP that was sent (assuming it's stored in a variable called otp)
        if entered_otp == otp:  # Replace 'otp' with the actual OTP sent
            messagebox.showinfo("Success", "OTP verification successful!")
        else:
            messagebox.showerror("Error", "Invalid OTP. Please try again.")
        

    sendotp = Button(frame4, text="Send\nOTP", bg="grey", fg="white", font=("Arial", 8, "bold"), padx=10,command=sendmail)
    sendotp.place(relx=0.83, rely=0.53)
    #VERIFY otp
    sendotp = Button(frame4, text="Verify", bg="grey", fg="white", font=("Arial", 8, "bold"), padx=10,command=verify_otp)
    sendotp.place(relx=0.83, rely=0.63)
    #otp>>>>>>>>>>>>>.............................................................................................
    # Create the Entry widget
    verify_otp_entry = Entry(frame4, font=("Arial", 12, "bold"), width=20, bd=0)
    verify_otp_entry.place(relx=0.3, rely=0.63)  # Place the Entry widget on the frame

    # Create a separating line
    Frame(frame4, width=150, height=2, bg='firebrick1').place(relx=0.3, rely=0.67)

    # verify_otp=Entry(frame4, font=("Arial", 12, "bold"),width=20,bd=0 ).place(relx=0.3,rely=0.63)
    # Frame(frame4, width=150, height=2, bg='firebrick1').place(relx=0.3, rely=0.67)
    # #next button
    # next_button2 = Button(frame4, text="Back", bg="blue", fg="white", font=("Arial", 16, "bold"), padx=20, command=back2)
    # next_button2.place(relx=0.1, rely=0.9)

    exit_button2 = Button(frame4, text="Submit", bg="blue", fg="white", font=("Arial", 16, "bold"), padx=15, command=validate_and_store_data)
    exit_button2.place(relx=0.54, rely=0.9)

    validate_button = Button(frame4, text="Verify", bg="red", fg="white", font=("Arial", 10, "bold"), padx=15, command=validate_captcha)
    validate_button.place(relx=0.6, rely=0.8)

    regenerate_button = Button(frame4, text="Regenerate", bg="red", fg="white", font=("Arial", 10, "bold"), padx=15, command=generate_and_draw_captcha)
    regenerate_button.place(relx=0.6, rely=0.72)

def student_ID_card():
    image6 = Label(image=bg_image6)
    image6.place(x=0, y=0, anchor="nw")

    def done(): 
        image6.destroy()
        rip.update()
        welcome_page()    

    #STUDENT NAME
    info=Entry(rip,bd=0,font=(12),width=10)
    info.insert(0,li[0])
    info.place(relx=0.4,rely=0.16)
    info.config(state='readonly',readonlybackground='white')
    #DOB
    info1=Entry(rip,bd=0,font=(12),width=10)
    info1.insert(0,li[1])
    info1.place(relx=0.4,rely=0.25)
    info1.config(state='readonly',readonlybackground='white')
    #ADDRESS
    info1=Entry(rip,bd=0,font=(12),width=10)
    info1.insert(0,li[3])
    info1.place(relx=0.68,rely=0.53)
    info1.config(state='readonly',readonlybackground='white')
    #ZIP CODE
    info1=Entry(rip,bd=0,font=(12),width=10)
    info1.insert(0,li[4])
    info1.place(relx=0.68,rely=0.64)
    info1.config(state='readonly',readonlybackground='white')
    #STATE
    info1=Entry(rip,bd=0,font=(12),width=10)
    info1.insert(0,li[5])
    info1.place(relx=0.4,rely=0.63)
    info1.config(state='readonly',readonlybackground='white')
    #fatherName
    info1=Entry(rip,bd=0,font=(12),width=10)
    info1.insert(0,li[6])
    info1.place(relx=0.4,rely=0.34)
    info1.config(state='readonly',readonlybackground='white')         
    #fatherName
    info1=Entry(rip,bd=0,font=(12),width=10)
    info1.insert(0,li[8])
    info1.place(relx=0.4,rely=0.43)
    info1.config(state='readonly',readonlybackground='white')
    #fatherName
    info1=Entry(rip,bd=0,font=(12),width=10)
    info1.insert(0,li[10])
    info1.place(relx=0.68,rely=0.34)
    info1.config(state='readonly',readonlybackground='white')
    #fatherName
    info1=Entry(rip,bd=0,font=(12),width=10)
    info1.insert(0,li[11])
    info1.place(relx=0.68,rely=0.43)
    info1.config(state='readonly',readonlybackground='white')
    #studentID
    info1=Entry(rip,bd=0,font=(12),width=10)
    info1.insert(0,li[13])
    info1.place(relx=0.68,rely=0.16)
    info1.config(state='readonly',readonlybackground='white')
    #CLASs
    info1=Entry(rip,bd=0,font=(12),width=10)
    info1.insert(0,li[14])
    info1.place(relx=0.68,rely=0.248)
    info1.config(state='readonly',readonlybackground='white')
    #CLASs
    info1=Entry(rip,bd=0,font=(12),width=15)
    info1.insert(0,li[16])
    info1.place(relx=0.39,rely=0.525)
    info1.config(state='readonly',readonlybackground='white')    


    Done = Button(text="Done", bg="Green", fg="white", font=("Arial", 16, "bold"), padx=20,command=done)
    Done.place(relx=0.48, rely=0.73)                         
welcome_page()
rip.mainloop()